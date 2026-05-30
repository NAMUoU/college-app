from flask import Flask, render_template, redirect, url_for, flash, request, send_file, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from datetime import datetime, date
import os
from openpyxl import Workbook

from config import Config
from models import db, User, StudentProfile, TeacherProfile, Group, Subject, Event, EventRegistration, CertificateRequest, Grade, Attendance, Notification, ActionLog
from forms import *
from utils import role_required, log_action, create_notification

app = Flask(__name__)
app.config.from_object(Config)

db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Пожалуйста, авторизуйтесь для доступа к этой странице'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.context_processor
def utility_processor():
    def get_unread_notifications_count():
        if current_user.is_authenticated:
            return Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
        return 0
    return dict(unread_count=get_unread_notifications_count())

@app.route('/')
def index():
    events = Event.query.order_by(Event.event_date).limit(4).all()
    return render_template('index.html', events=events)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user and user.check_password(form.password.data):
            if not user.is_active:
                flash('Ваша учетная запись заблокирована', 'danger')
                return redirect(url_for('login'))
            
            login_user(user)
            log_action('Вход в систему')
            
            if user.must_change_password:
                flash('Необходимо сменить пароль при первом входе', 'warning')
                return redirect(url_for('profile'))
            
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            flash('Неверный email или пароль', 'danger')
    
    return render_template('login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    log_action('Выход из системы')
    logout_user()
    flash('Вы вышли из системы', 'info')
    return redirect(url_for('index'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = RegistrationForm()
    form.group_id.choices = [(0, 'Выберите группу')] + [(g.id, g.name) for g in Group.query.all()]
    
    if form.validate_on_submit():
        if User.query.filter_by(email=form.email.data).first():
            flash('Пользователь с таким email уже существует', 'danger')
            return redirect(url_for('register'))
        
        user = User(
            email=form.email.data,
            role='student',
            is_active=True
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        
        student = StudentProfile(
            user_id=user.id,
            full_name=form.full_name.data,
            group_id=form.group_id.data if form.group_id.data != 0 else None,
            phone=form.phone.data,
            status='pending'
        )
        db.session.add(student)
        db.session.commit()
        
        log_action(f'Регистрация студента {form.full_name.data}')
        create_notification(user.id, 'Заявка на регистрацию', 'Ваша заявка на регистрацию отправлена администратору на рассмотрение')
        
        flash('Регистрация успешна! Ожидайте подтверждения администратора', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html', form=form)

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    form = ForgotPasswordForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user:
            flash('Инструкции по восстановлению пароля отправлены на ваш email (в учебной версии проверьте консоль)', 'info')
            print(f'[DEMO] Восстановление пароля для {user.email}')
        else:
            flash('Пользователь с таким email не найден', 'danger')
    return render_template('forgot_password.html', form=form)

@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.role == 'student':
        student = current_user.student_profile
        if student.status != 'approved':
            return render_template('dashboard.html', student=student)
        
        recent_grades = Grade.query.filter_by(student_id=student.id).order_by(Grade.grade_date.desc()).limit(5).all()
        upcoming_events = EventRegistration.query.filter_by(student_id=student.id).join(Event).filter(Event.event_date >= date.today()).order_by(Event.event_date).limit(3).all()
        recent_notifications = Notification.query.filter_by(user_id=current_user.id, is_read=False).order_by(Notification.created_at.desc()).limit(5).all()
        
        subjects_grades = {}
        for grade in Grade.query.filter_by(student_id=student.id).all():
            if grade.subject_id not in subjects_grades:
                subjects_grades[grade.subject_id] = []
            subjects_grades[grade.subject_id].append(grade.grade_value)
        
        avg_grades = {sid: sum(grades)/len(grades) for sid, grades in subjects_grades.items()}
        
        return render_template('dashboard.html', student=student, recent_grades=recent_grades, 
                             upcoming_events=upcoming_events, recent_notifications=recent_notifications, avg_grades=avg_grades)
    
    elif current_user.role == 'teacher':
        teacher = current_user.teacher_profile
        subjects = Subject.query.filter_by(teacher_id=teacher.id).all()
        students_count = StudentProfile.query.filter_by(status='approved').count()
        grades_count = Grade.query.filter_by(teacher_id=teacher.id).count()
        
        return render_template('dashboard.html', teacher=teacher, subjects=subjects, 
                             students_count=students_count, grades_count=grades_count)
    
    elif current_user.role == 'admin':
        students_count = StudentProfile.query.count()
        teachers_count = TeacherProfile.query.count()
        events_count = Event.query.count()
        pending_requests = StudentProfile.query.filter_by(status='pending').count()
        
        return render_template('dashboard.html', students_count=students_count, 
                             teachers_count=teachers_count, events_count=events_count, pending_requests=pending_requests)
    
    return render_template('dashboard.html')

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    form = ProfileForm()
    
    if current_user.role == 'student':
        student = current_user.student_profile
        if form.validate_on_submit():
            student.full_name = form.full_name.data
            student.phone = form.phone.data
            current_user.email = form.email.data
            
            if form.current_password.data and form.new_password.data:
                if current_user.check_password(form.current_password.data):
                    current_user.set_password(form.new_password.data)
                    flash('Пароль успешно изменен', 'success')
                    log_action('Смена пароля')
                else:
                    flash('Неверный текущий пароль', 'danger')
                    return redirect(url_for('profile'))
            
            db.session.commit()
            log_action('Обновление профиля')
            flash('Профиль успешно обновлен', 'success')
            return redirect(url_for('profile'))
        
        form.full_name.data = student.full_name
        form.phone.data = student.phone
        form.email.data = current_user.email
    
    elif current_user.role == 'teacher':
        teacher = current_user.teacher_profile
        if form.validate_on_submit():
            teacher.full_name = form.full_name.data
            current_user.email = form.email.data
            
            if form.current_password.data and form.new_password.data:
                if current_user.check_password(form.current_password.data):
                    current_user.set_password(form.new_password.data)
                    current_user.must_change_password = False
                    flash('Пароль успешно изменен', 'success')
                    log_action('Смена пароля')
                else:
                    flash('Неверный текущий пароль', 'danger')
                    return redirect(url_for('profile'))
            
            db.session.commit()
            log_action('Обновление профиля')
            flash('Профиль успешно обновлен', 'success')
            return redirect(url_for('profile'))
        
        form.full_name.data = teacher.full_name
        form.email.data = current_user.email
    
    return render_template('profile.html', form=form)

# Студенческие маршруты
@app.route('/student/events')
@login_required
@role_required('student')
def student_events():
    student = current_user.student_profile
    if student.status != 'approved':
        flash('Ваша учетная запись еще не подтверждена', 'warning')
        return redirect(url_for('dashboard'))
    
    events = Event.query.order_by(Event.event_date).all()
    registered_events = [reg.event_id for reg in EventRegistration.query.filter_by(student_id=student.id).all()]
    
    return render_template('student/events.html', events=events, registered_events=registered_events)

@app.route('/student/register_event/<int:event_id>')
@login_required
@role_required('student')
def register_event(event_id):
    student = current_user.student_profile
    event = Event.query.get_or_404(event_id)
    
    existing = EventRegistration.query.filter_by(event_id=event_id, student_id=student.id).first()
    if existing:
        flash('Вы уже записаны на это мероприятие', 'warning')
    else:
        registration = EventRegistration(event_id=event_id, student_id=student.id)
        db.session.add(registration)
        db.session.commit()
        log_action(f'Запись на мероприятие {event.title}')
        create_notification(current_user.id, 'Запись на мероприятие', f'Вы записались на мероприятие "{event.title}"')
        flash('Вы успешно записались на мероприятие', 'success')
    
    return redirect(url_for('student_events'))

@app.route('/student/my_events')
@login_required
@role_required('student')
def student_my_events():
    student = current_user.student_profile
    registrations = EventRegistration.query.filter_by(student_id=student.id).join(Event).order_by(Event.event_date).all()
    return render_template('student/my_events.html', registrations=registrations)

@app.route('/student/cancel_event/<int:registration_id>')
@login_required
@role_required('student')
def cancel_event(registration_id):
    registration = EventRegistration.query.get_or_404(registration_id)
    if registration.student_id != current_user.student_profile.id:
        flash('Нет доступа', 'danger')
        return redirect(url_for('student_my_events'))
    
    event_title = registration.event.title
    db.session.delete(registration)
    db.session.commit()
    log_action(f'Отмена записи на мероприятие {event_title}')
    create_notification(current_user.id, 'Отмена записи', f'Вы отменили запись на мероприятие "{event_title}"')
    flash('Запись отменена', 'success')
    return redirect(url_for('student_my_events'))

@app.route('/student/certificates')
@login_required
@role_required('student')
def student_certificates():
    student = current_user.student_profile
    requests = CertificateRequest.query.filter_by(student_id=student.id).order_by(CertificateRequest.created_at.desc()).all()
    return render_template('student/certificates.html', requests=requests)

@app.route('/student/request_certificate')
@login_required
@role_required('student')
def request_certificate():
    student = current_user.student_profile
    request = CertificateRequest(student_id=student.id, status='pending')  # ← Ошибка здесь
    db.session.add(request)
    db.session.commit()
    log_action('Заказ справки об обучении')
    create_notification(current_user.id, 'Заказ справки', 'Ваша заявка на справку отправлена в обработку')
    flash('Заявка на справку успешно отправлена', 'success')
    return redirect(url_for('student_certificates'))

@app.route('/student/grades')
@login_required
@role_required('student')
def student_grades():
    student = current_user.student_profile
    grades = Grade.query.filter_by(student_id=student.id).join(Subject).order_by(Grade.grade_date.desc()).all()
    
    subjects_grades = {}
    for grade in grades:
        if grade.subject_id not in subjects_grades:
            subjects_grades[grade.subject_id] = {'subject': grade.subject.name, 'grades': []}
        subjects_grades[grade.subject_id]['grades'].append(grade)
    
    for subject_id in subjects_grades:
        grades_list = [g.grade_value for g in subjects_grades[subject_id]['grades']]
        subjects_grades[subject_id]['average'] = sum(grades_list) / len(grades_list) if grades_list else 0
    
    return render_template('student/grades.html', subjects_grades=subjects_grades)

@app.route('/student/notifications')
@login_required
@role_required('student')
def student_notifications():
    notifications = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.created_at.desc()).all()
    for notif in notifications:
        if not notif.is_read:
            notif.is_read = True
    db.session.commit()
    return render_template('student/notifications.html', notifications=notifications)

@app.route('/download_schedule')
@login_required
def download_schedule():
    file_path = os.path.join(app.static_folder, 'files', 'schedule.xlsx')
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Расписание"
        ws['A1'] = "Расписание занятий"
        wb.save(file_path)
    
    return send_file(file_path, as_attachment=True)

# Преподавательские маршруты
@app.route('/teacher/journal', methods=['GET', 'POST'])
@login_required
@role_required('teacher')
def teacher_journal():
    teacher = current_user.teacher_profile
    form = JournalSelectForm()
    
    form.group_id.choices = [(0, 'Выберите группу')] + [(g.id, g.name) for g in Group.query.all()]
    form.subject_id.choices = [(0, 'Выберите дисциплину')] + [(s.id, s.name) for s in Subject.query.filter_by(teacher_id=teacher.id).all()]
    
    students = []
    selected_group = None
    selected_subject = None
    selected_date = date.today()
    
    if request.method == 'POST' and form.validate_on_submit():
        selected_group = Group.query.get(form.group_id.data)
        selected_subject = Subject.query.get(form.subject_id.data)
        selected_date = form.lesson_date.data
        
        if selected_group and selected_subject and selected_group.id != 0 and selected_subject.id != 0:
            students = StudentProfile.query.filter_by(group_id=selected_group.id, status='approved').all()
    
    if students and selected_subject and selected_date:
        attendances = {att.student_id: att for att in Attendance.query.filter_by(
            subject_id=selected_subject.id, lesson_date=selected_date
        ).all()}
        
        grades_for_date = {grade.student_id: grade for grade in Grade.query.filter_by(
            subject_id=selected_subject.id, grade_date=selected_date, teacher_id=teacher.id
        ).all()}
        
        return render_template('teacher/journal.html', students=students, selected_group=selected_group,
                             selected_subject=selected_subject, selected_date=selected_date,
                             attendances=attendances, grades_for_date=grades_for_date, form=form)
    
    return render_template('teacher/journal.html', form=form, students=students)

@app.route('/teacher/save_attendance', methods=['POST'])
@login_required
@role_required('teacher')
def save_attendance():
    teacher = current_user.teacher_profile
    student_id = request.form.get('student_id')
    subject_id = request.form.get('subject_id')
    lesson_date = datetime.strptime(request.form.get('lesson_date'), '%Y-%m-%d').date()
    status = request.form.get('status')
    
    attendance = Attendance.query.filter_by(
        student_id=student_id, subject_id=subject_id, lesson_date=lesson_date    ).first()
    
    if attendance:
        attendance.status = status
        attendance.teacher_id = teacher.id
    else:
        attendance = Attendance(
            student_id=student_id, subject_id=subject_id, teacher_id=teacher.id,
            lesson_date=lesson_date, status=status
        )
        db.session.add(attendance)
    
    db.session.commit()
    log_action(f'Отметка посещаемости студента ID {student_id}')
    return jsonify({'success': True})

@app.route('/teacher/save_grade', methods=['POST'])
@login_required
@role_required('teacher')
def save_grade():
    teacher = current_user.teacher_profile
    student_id = request.form.get('student_id')
    subject_id = request.form.get('subject_id')
    lesson_date = datetime.strptime(request.form.get('lesson_date'), '%Y-%m-%d').date()
    grade_value = int(request.form.get('grade_value'))
    comment = request.form.get('comment', '')
    
    grade = Grade.query.filter_by(
        student_id=student_id, subject_id=subject_id, grade_date=lesson_date, teacher_id=teacher.id
    ).first()
    
    if grade:
        grade.grade_value = grade_value
        grade.comment = comment
    else:
        grade = Grade(
            student_id=student_id, subject_id=subject_id, teacher_id=teacher.id,
            grade_date=lesson_date, grade_value=grade_value, comment=comment
        )
        db.session.add(grade)
        
        student = StudentProfile.query.get(student_id)
        if student and student.user:
            subject = Subject.query.get(subject_id)
            create_notification(student.user_id, 'Новая оценка', 
                              f'По предмету "{subject.name if subject else 'предмет'}" выставлена оценка {grade_value}')
    
    db.session.commit()
    log_action(f'Выставление оценки студенту ID {student_id}')
    return jsonify({'success': True})

@app.route('/teacher/export_journal')
@login_required
@role_required('teacher')
def export_journal():
    teacher = current_user.teacher_profile
    group_id = request.args.get('group_id', type=int)
    subject_id = request.args.get('subject_id', type=int)
    
    if not group_id or not subject_id:
        flash('Необходимо выбрать группу и дисциплину', 'danger')
        return redirect(url_for('teacher_journal'))
    
    group = Group.query.get(group_id)
    subject = Subject.query.get(subject_id)
    students = StudentProfile.query.filter_by(group_id=group_id, status='approved').all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Журнал {group.name} {subject.name}"
    
    ws['A1'] = f"Журнал успеваемости группы {group.name} по дисциплине {subject.name}"
    ws['A2'] = f"Преподаватель: {teacher.full_name}"
    ws['A3'] = "Дата генерации: " + datetime.now().strftime("%d.%m.%Y %H:%M")
    
    headers = ['№', 'ФИО студента']
    all_dates = set()
    
    for student in students:
        grades = Grade.query.filter_by(student_id=student.id, subject_id=subject_id).all()
        for grade in grades:
            all_dates.add(grade.grade_date)
    
    sorted_dates = sorted(all_dates)
    for d in sorted_dates:
        headers.append(d.strftime("%d.%m.%Y"))
    headers.append('Средний балл')
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=header)
    
    for row, student in enumerate(students, start=6):
        ws.cell(row=row, column=1, value=row-5)
        ws.cell(row=row, column=2, value=student.full_name)
        
        grades_dict = {g.grade_date: g.grade_value for g in Grade.query.filter_by(student_id=student.id, subject_id=subject_id).all()}
        
        total = 0
        count = 0
        for col, d in enumerate(sorted_dates, start=3):
            if d in grades_dict:
                ws.cell(row=row, column=col, value=grades_dict[d])
                total += grades_dict[d]
                count += 1
            else:
                ws.cell(row=row, column=col, value='')
        
        average = total / count if count > 0 else 0
        ws.cell(row=row, column=len(headers), value=round(average, 2))
    
    file_path = os.path.join(app.static_folder, f'journal_{group.id}_{subject.id}.xlsx')
    wb.save(file_path)
    
    return send_file(file_path, as_attachment=True, download_name=f'journal_{group.name}_{subject.name}.xlsx')

@app.route('/teacher/grades')
@login_required
@role_required('teacher')
def teacher_grades():
    return render_template('teacher/grades_list.html')

@app.route('/teacher/export')
@login_required
@role_required('teacher')
def teacher_export():
    groups = Group.query.all()
    return render_template('teacher/export.html', groups=groups)

# Администраторские маршруты
@app.route('/admin/students')
@login_required
@role_required('admin')
def admin_students():
    students = StudentProfile.query.all()
    groups = Group.query.all()
    return render_template('admin/students.html', students=students, groups=groups)

@app.route('/admin/student_requests')
@login_required
@role_required('admin')
def admin_student_requests():
    pending_students = StudentProfile.query.filter_by(status='pending').all()
    return render_template('admin/student_requests.html', pending_students=pending_students)

@app.route('/admin/approve_student/<int:student_id>')
@login_required
@role_required('admin')
def approve_student(student_id):
    student = StudentProfile.query.get_or_404(student_id)
    student.status = 'approved'
    db.session.commit()
    log_action(f'Подтверждение студента {student.full_name}')
    create_notification(student.user_id, 'Заявка одобрена', 'Ваша заявка на регистрацию одобрена администратором')
    flash('Студент подтвержден', 'success')
    return redirect(url_for('admin_student_requests'))

@app.route('/admin/reject_student/<int:student_id>')
@login_required
@role_required('admin')
def reject_student(student_id):
    student = StudentProfile.query.get_or_404(student_id)
    student.status = 'rejected'
    db.session.commit()
    log_action(f'Отклонение студента {student.full_name}')
    create_notification(student.user_id, 'Заявка отклонена', 'Ваша заявка на регистрацию отклонена администратором')
    flash('Студент отклонен', 'success')
    return redirect(url_for('admin_student_requests'))

@app.route('/admin/toggle_student_block/<int:student_id>')
@login_required
@role_required('admin')
def toggle_student_block(student_id):
    student = StudentProfile.query.get_or_404(student_id)
    if student.status == 'blocked':
        student.status = 'approved'
        student.user.is_active = True
        flash('Студент разблокирован', 'success')
    else:
        student.status = 'blocked'
        student.user.is_active = False
        flash('Студент заблокирован', 'success')
    db.session.commit()
    return redirect(request.referrer or url_for('admin_students'))

@app.route('/admin/teachers')
@login_required
@role_required('admin')
def admin_teachers():
    teachers = TeacherProfile.query.all()
    return render_template('admin/teachers.html', teachers=teachers)

@app.route('/admin/add_teacher', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def add_teacher():
    form = TeacherForm()
    
    if form.validate_on_submit():
        if User.query.filter_by(email=form.email.data).first():
            flash('Пользователь с таким email уже существует', 'danger')
            return redirect(url_for('add_teacher'))
        
        user = User(email=form.email.data, role='teacher', is_active=True, must_change_password=True)
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        
        teacher = TeacherProfile(user_id=user.id, full_name=form.full_name.data)
        db.session.add(teacher)
        db.session.commit()
        
        log_action(f'Создание преподавателя {form.full_name.data}')
        flash('Преподаватель успешно добавлен', 'success')
        return redirect(url_for('admin_teachers'))
    
    return render_template('admin/add_teacher.html', form=form)

@app.route('/admin/delete_teacher/<int:teacher_id>')
@login_required
@role_required('admin')
def delete_teacher(teacher_id):
    teacher = TeacherProfile.query.get_or_404(teacher_id)
    user = teacher.user
    db.session.delete(teacher)
    db.session.delete(user)
    db.session.commit()
    log_action(f'Удаление преподавателя {teacher.full_name}')
    flash('Преподаватель удален', 'success')
    return redirect(url_for('admin_teachers'))

@app.route('/admin/groups')
@login_required
@role_required('admin')
def admin_groups():
    groups = Group.query.all()
    return render_template('admin/groups.html', groups=groups)

@app.route('/admin/add_group', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def add_group():
    form = GroupForm()
    if form.validate_on_submit():
        if Group.query.filter_by(name=form.name.data).first():
            flash('Группа с таким названием уже существует', 'danger')
            return redirect(url_for('add_group'))
        
        group = Group(name=form.name.data)
        db.session.add(group)
        db.session.commit()
        log_action(f'Создание группы {form.name.data}')
        flash('Группа успешно добавлена', 'success')
        return redirect(url_for('admin_groups'))
    return render_template('admin/add_group.html', form=form)

@app.route('/admin/delete_group/<int:group_id>')
@login_required
@role_required('admin')
def delete_group(group_id):
    group = Group.query.get_or_404(group_id)
    db.session.delete(group)
    db.session.commit()
    log_action(f'Удаление группы {group.name}')
    flash('Группа удалена', 'success')
    return redirect(url_for('admin_groups'))

@app.route('/admin/subjects')
@login_required
@role_required('admin')
def admin_subjects():
    subjects = Subject.query.all()
    teachers = TeacherProfile.query.all()
    return render_template('admin/subjects.html', subjects=subjects, teachers=teachers)

@app.route('/admin/add_subject', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def add_subject():
    form = SubjectForm()
    form.teacher_id.choices = [(0, 'Выберите преподавателя')] + [(t.id, t.full_name) for t in TeacherProfile.query.all()]
    
    if form.validate_on_submit():
        subject = Subject(name=form.name.data, teacher_id=form.teacher_id.data if form.teacher_id.data != 0 else None)
        db.session.add(subject)
        db.session.commit()
        log_action(f'Создание дисциплины {form.name.data}')
        flash('Дисциплина успешно добавлена', 'success')
        return redirect(url_for('admin_subjects'))
    
    return render_template('admin/add_subject.html', form=form)

@app.route('/admin/delete_subject/<int:subject_id>')
@login_required
@role_required('admin')
def delete_subject(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    db.session.delete(subject)
    db.session.commit()
    log_action(f'Удаление дисциплины {subject.name}')
    flash('Дисциплина удалена', 'success')
    return redirect(url_for('admin_subjects'))

@app.route('/admin/events')
@login_required
@role_required('admin')
def admin_events():
    events = Event.query.all()
    return render_template('admin/events.html', events=events)

@app.route('/admin/add_event', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def add_event():
    form = EventForm()
    if form.validate_on_submit():
        event = Event(
            title=form.title.data,
            event_date=form.event_date.data,
            event_time=form.event_time.data,
            place=form.place.data,
            description=form.description.data
        )
        db.session.add(event)
        db.session.commit()
        log_action(f'Создание мероприятия {form.title.data}')
        flash('Мероприятие успешно добавлено', 'success')
        return redirect(url_for('admin_events'))
    return render_template('admin/add_event.html', form=form)

@app.route('/admin/delete_event/<int:event_id>')
@login_required
@role_required('admin')
def delete_event(event_id):
    event = Event.query.get_or_404(event_id)
    db.session.delete(event)
    db.session.commit()
    log_action(f'Удаление мероприятия {event.title}')
    flash('Мероприятие удалено', 'success')
    return redirect(url_for('admin_events'))

@app.route('/admin/certificates')
@login_required
@role_required('admin')
def admin_certificates():
    requests_list = CertificateRequest.query.all()
    return render_template('admin/certificates.html', requests=requests_list)

@app.route('/admin/update_certificate/<int:request_id>', methods=['POST'])
@login_required
@role_required('admin')
def update_certificate(request_id):
    cert_request = CertificateRequest.query.get_or_404(request_id)
    new_status = request.form.get('status')
    cert_request.status = new_status
    db.session.commit()
    log_action(f'Изменение статуса справки на {new_status}')
    if cert_request.student and cert_request.student.user:
        create_notification(cert_request.student.user_id, 'Статус справки изменен', f'Статус вашей справки: {new_status}')
    flash('Статус справки обновлен', 'success')
    return redirect(url_for('admin_certificates'))

@app.route('/admin/logs')
@login_required
@role_required('admin')
def admin_logs():
    logs = ActionLog.query.order_by(ActionLog.created_at.desc()).limit(100).all()
    return render_template('admin/logs.html', logs=logs)

@app.route('/admin/reports')
@login_required
@role_required('admin')
def admin_reports():
    active_students = StudentProfile.query.filter_by(status='approved').count()
    teachers_count = TeacherProfile.query.count()
    events_count = Event.query.count()
    
    events_popularity = []
    for event in Event.query.all():
        count = EventRegistration.query.filter_by(event_id=event.id).count()
        events_popularity.append({'name': event.title, 'count': count})
    events_popularity.sort(key=lambda x: x['count'], reverse=True)
    
    teacher_load = []
    for teacher in TeacherProfile.query.all():
        subjects_count = Subject.query.filter_by(teacher_id=teacher.id).count()
        teacher_load.append({'name': teacher.full_name, 'subjects': subjects_count})
    
    return render_template('admin/reports.html', active_students=active_students, 
                         teachers_count=teachers_count, events_count=events_count,
                         events_popularity=events_popularity[:5], teacher_load=teacher_load)

@app.route('/admin/event_participants/<int:event_id>')
@login_required
@role_required('admin')
def event_participants(event_id):
    event = Event.query.get_or_404(event_id)
    registrations = EventRegistration.query.filter_by(event_id=event_id).join(StudentProfile).all()
    
    if request.args.get('export') == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = f"Участники {event.title}"
        ws['A1'] = f"Список участников мероприятия: {event.title}"
        ws['A2'] = f"Дата: {event.event_date} {event.event_time}"
        ws['A3'] = f"Место: {event.place}"
        ws['A5'] = "№"
        ws['B5'] = "ФИО"
        ws['C5'] = "Группа"
        ws['D5'] = "Телефон"
        ws['E5'] = "Email"
        
        for idx, reg in enumerate(registrations, start=6):
            ws.cell(row=idx, column=1, value=idx-5)
            ws.cell(row=idx, column=2, value=reg.student.full_name)
            ws.cell(row=idx, column=3, value=reg.student.group.name if reg.student.group else 'Нет группы')
            ws.cell(row=idx, column=4, value=reg.student.phone)
            ws.cell(row=idx, column=5, value=reg.student.user.email)
        
        file_path = os.path.join(app.static_folder, f'event_{event_id}_participants.xlsx')
        wb.save(file_path)
        return send_file(file_path, as_attachment=True, download_name=f'participants_{event.title}.xlsx')
    
    return render_template('admin/event_participants.html', event=event, registrations=registrations)

@app.route('/admin/grades')
@login_required
@role_required('admin')
def admin_grades():
    students = StudentProfile.query.filter_by(status='approved').all()
    subjects = Subject.query.all()
    groups = Group.query.all()
    
    selected_group_id = request.args.get('group_id', type=int)
    selected_subject_id = request.args.get('subject_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    grades_query = Grade.query
    
    if selected_group_id:
        students_in_group = StudentProfile.query.filter_by(group_id=selected_group_id).all()
        student_ids = [s.id for s in students_in_group]
        grades_query = grades_query.filter(Grade.student_id.in_(student_ids))
    
    if selected_subject_id:
        grades_query = grades_query.filter_by(subject_id=selected_subject_id)
    
    if start_date:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        grades_query = grades_query.filter(Grade.grade_date >= start_date_obj)
    
    if end_date:
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        grades_query = grades_query.filter(Grade.grade_date <= end_date_obj)
    
    grades = grades_query.order_by(Grade.grade_date.desc()).all()
    
    return render_template('admin/grades.html', grades=grades, students=students, 
                         subjects=subjects, groups=groups,
                         selected_group_id=selected_group_id, selected_subject_id=selected_subject_id,
                         start_date=start_date, end_date=end_date)

@app.route('/admin/update_student_group/<int:student_id>', methods=['POST'])
@login_required
@role_required('admin')
def update_student_group(student_id):
    student = StudentProfile.query.get_or_404(student_id)
    new_group_id = request.form.get('group_id')
    
    if new_group_id and new_group_id != '0':
        student.group_id = int(new_group_id)
        db.session.commit()
        log_action(f'Изменение группы студента {student.full_name}')
        flash('Группа студента обновлена', 'success')
    else:
        flash('Выберите группу', 'danger')
    
    return redirect(url_for('admin_students'))

@app.route('/admin/delete_student/<int:student_id>')
@login_required
@role_required('admin')

def delete_student(student_id):
    student = StudentProfile.query.get_or_404(student_id)
    user = student.user
    
    # Сохраняем имя для лога
    student_name = student.full_name
    
    try:
        # Удаляем все связанные данные (каскадно)
        # Сначала удаляем регистрации на мероприятия
        for reg in student.event_registrations.all():
            db.session.delete(reg)
        
        # Удаляем заявки на справки
        for cert in student.certificate_requests.all():
            db.session.delete(cert)
        
        # Удаляем оценки
        for grade in student.grades.all():
            db.session.delete(grade)
        
        # Удаляем посещаемость
        for att in student.attendances.all():
            db.session.delete(att)
        
        # Удаляем профиль студента
        db.session.delete(student)
        
        # Удаляем пользователя
        db.session.delete(user)
        
        db.session.commit()
        
        log_action(f'Полное удаление студента {student_name} (ID: {student_id})')
        flash(f'Студент {student_name} полностью удален из системы', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при удалении: {str(e)}', 'danger')
        print(f'Error deleting student: {e}')
    
    return redirect(url_for('admin_students'))

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        print("✅ Database tables created/verified")
    app.run(debug=True)